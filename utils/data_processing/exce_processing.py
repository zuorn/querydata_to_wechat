import pandas as pd
from openpyxl.styles import PatternFill, Border, Side
from openpyxl import load_workbook
from utils.logs import log



def exce_processing(file_path):
    """
    excel文件处理函数，包括：
    1. 冻结首行
    2. 将列名填充单元格颜色为绿色
    3. 设置自适应列宽
    4. 设置所有数据单元格设置所有框线
    """
    # 读取Excel文件
    df = pd.read_excel(file_path)
    
    # 冻结首行
    df.to_excel(file_path, index=False, header=True, freeze_panes=(1,1))
    print('🔧冻结首行成功')
    log.info("🔧冻结首行成功")
    

    # 加载工作簿
    wb = load_workbook(file_path)
    ws = wb.active

    # 将列名填充单元格颜色为绿色
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='lightUp')
    print('🔧列名填充单元格颜色为绿色成功')
    log.info("🔧列名填充单元格颜色为绿色成功")

    # 设置自适应列宽
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    print('🔧设置自适应列宽成功')
    log.info("🔧设置自适应列宽成功")

    # 设置所有数据单元格的框线
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
    print('🔧设置所有数据单元格的框线成功')
    log.info("🔧设置所有数据单元格的框线成功")

    wb.save(file_path)

    def exce_processing_config():
        """
        配置处理excel文件函数，包括：
        1. 
        """

# 测试
if __name__ == '__main__':
    exce_processing('output/任务2.xlsx')